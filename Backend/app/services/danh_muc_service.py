from datetime import datetime
from decimal import Decimal
from io import BytesIO
from uuid import uuid4
from zoneinfo import ZoneInfo

from fastapi import HTTPException, UploadFile, status
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from sqlalchemy.exc import IntegrityError
from sqlmodel import Session
from starlette.responses import StreamingResponse

from app.core.rbac import ADMIN_ROLE_IDS, HCNS_ROLE_IDS
from app.models.lich_hop import LichHop
from app.models.phong_hop import PhongHop
from app.models.quy_phep import QuyPhep
from app.models.quyen import Quyen
from app.models.tai_san import TaiSan
from app.models.thanh_vien_lich_hop import ThanhVienLichHop
from app.models.tai_khoan import TaiKhoan
from app.models.thong_bao import ThongBao
from app.repositories.danh_muc_repo import GenericDanhMucRepository, LichHopRepository, PhongHopRepository
from app.schemas.danh_muc_schema import (
    LichHopCreate,
    LichHopResponse,
    LichHopXuLyRequest,
    NhanVienOptionResponse,
    PhongHopCreate,
    PhongHopUpdate,
    QuyenCreate,
    QuyenUpdate,
    QuyPhepCreate,
    QuyPhepUpdate,
    TaiSanCreate,
    TaiSanUpdate,
    ThanhVienLichHopResponse,
)

MANAGER_ROLE_IDS = {"MANAGER", "Manager", "QUAN_LY", "QuanLy"}
MEETING_MODERATOR_ROLE_IDS = ADMIN_ROLE_IDS | HCNS_ROLE_IDS | MANAGER_ROLE_IDS
APP_TIMEZONE = ZoneInfo("Asia/Ho_Chi_Minh")


class DanhMucService:
    def __init__(self, session: Session) -> None:
        self.phong_hop_repo = PhongHopRepository(session)
        self.lich_hop_repo = LichHopRepository(session)
        self.tai_san_repo = GenericDanhMucRepository(session, TaiSan)
        self.quyen_repo = GenericDanhMucRepository(session, Quyen)
        self.quy_phep_repo = GenericDanhMucRepository(session, QuyPhep)

    def list_phong_hop(self, page: int = 1, size: int = 20) -> list[PhongHop]:
        return self.phong_hop_repo.list_paginated(page, size)

    def create_phong_hop(self, payload: PhongHopCreate) -> PhongHop:
        if self.phong_hop_repo.get(payload.id_Phong):
            raise HTTPException(status.HTTP_409_CONFLICT, "Ma phong hop da ton tai")
        return self.phong_hop_repo.create(PhongHop(**payload.model_dump()))

    def update_phong_hop(self, id_phong: str, payload: PhongHopUpdate) -> PhongHop:
        phong = self._get_phong_hop(id_phong)
        data = payload.model_dump(exclude_unset=True)
        data.pop("id_Phong", None)
        for key, value in data.items():
            setattr(phong, key, value)
        return self.phong_hop_repo.save(phong)

    def delete_phong_hop(self, id_phong: str) -> dict[str, str]:
        phong = self._get_phong_hop(id_phong)
        if self.phong_hop_repo.has_lich_hop_reference(id_phong):
            phong.trangThai = 0
            self.phong_hop_repo.save(phong)
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                "Phong hop dang duoc tham chieu, da chuyen sang trang thai ngung hoat dong",
            )
        self.phong_hop_repo.delete(phong)
        return {"id_Phong": id_phong}

    async def import_phong_hop(self, file: UploadFile) -> dict[str, object]:
        if not file.filename or not file.filename.lower().endswith(".xlsx"):
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Chi ho tro file Excel .xlsx")
        workbook = load_workbook(BytesIO(await file.read()))
        sheet = workbook.active
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        errors: list[dict[str, object]] = []
        success = 0
        for index, row in enumerate(rows, start=2):
            try:
                payload = PhongHopCreate(
                    id_Phong=str(row[0]),
                    tenPhong=str(row[1]),
                    sucChua=int(row[2]),
                    trangThai=int(row[3] if row[3] is not None else 1),
                    moTa=str(row[4]) if row[4] is not None else None,
                )
                self.create_phong_hop(payload)
                success += 1
            except Exception as exc:
                errors.append({"dong": index, "loi": str(exc)})
        return {
            "tong_dong": len(rows),
            "thanh_cong": success,
            "that_bai": len(errors),
            "chi_tiet_loi": errors,
        }

    def export_phong_hop(self) -> StreamingResponse:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "PhongHop"
        sheet.append(["id_Phong", "tenPhong", "sucChua", "trangThai", "moTa"])
        for item in self.phong_hop_repo.list_all(limit=10000):
            sheet.append([item.id_Phong, item.tenPhong, item.sucChua, item.trangThai, item.moTa])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=phong_hop.xlsx"},
        )

    def list_tai_san(self, page: int = 1, size: int = 20) -> list[TaiSan]:
        return self.tai_san_repo.list_paginated(page, size)

    def create_tai_san(self, payload: TaiSanCreate) -> TaiSan:
        return self._create_catalog_item(self.tai_san_repo, TaiSan(**payload.model_dump()), payload.id_TaiSan)

    def update_tai_san(self, id_tai_san: str, payload: TaiSanUpdate) -> TaiSan:
        return self._update_catalog_item(self.tai_san_repo, id_tai_san, payload)

    def delete_tai_san(self, id_tai_san: str) -> dict[str, str]:
        return self._delete_catalog_item(self.tai_san_repo, id_tai_san)

    async def import_tai_san(self, file: UploadFile) -> dict[str, object]:
        return await self._import_catalog(
            file=file,
            factory=lambda row: TaiSanCreate(
                id_TaiSan=str(row[0]),
                tenTaiSan=str(row[1]),
                serialNumber=str(row[2]) if row[2] is not None else None,
                ngayMua=row[3],
                giaTri=Decimal(str(row[4])) if row[4] is not None else None,
                tinhTrang=str(row[5]),
                trangThai=int(row[6] if row[6] is not None else 1),
            ),
            creator=self.create_tai_san,
        )

    def export_tai_san(self) -> StreamingResponse:
        return self._export_catalog(
            filename="tai_san.xlsx",
            sheet_name="TaiSan",
            headers=["id_TaiSan", "tenTaiSan", "serialNumber", "ngayMua", "giaTri", "tinhTrang", "trangThai"],
            rows=[
                [item.id_TaiSan, item.tenTaiSan, item.serialNumber, item.ngayMua, item.giaTri, item.tinhTrang, item.trangThai]
                for item in self.tai_san_repo.list_all(limit=10000)
            ],
        )

    def list_quyen(self, page: int = 1, size: int = 20) -> list[Quyen]:
        return self.quyen_repo.list_paginated(page, size)

    def create_quyen(self, payload: QuyenCreate) -> Quyen:
        return self._create_catalog_item(self.quyen_repo, Quyen(**payload.model_dump()), payload.id_Quyen)

    def update_quyen(self, id_quyen: str, payload: QuyenUpdate) -> Quyen:
        return self._update_catalog_item(self.quyen_repo, id_quyen, payload)

    def delete_quyen(self, id_quyen: str) -> dict[str, str]:
        return self._delete_catalog_item(self.quyen_repo, id_quyen)

    async def import_quyen(self, file: UploadFile) -> dict[str, object]:
        return await self._import_catalog(
            file=file,
            factory=lambda row: QuyenCreate(
                id_Quyen=str(row[0]),
                tenQuyen=str(row[1]),
                hanhDong=str(row[2]),
                moTa=str(row[3]) if row[3] is not None else None,
            ),
            creator=self.create_quyen,
        )

    def export_quyen(self) -> StreamingResponse:
        return self._export_catalog(
            filename="quyen.xlsx",
            sheet_name="Quyen",
            headers=["id_Quyen", "tenQuyen", "hanhDong", "moTa"],
            rows=[[item.id_Quyen, item.tenQuyen, item.hanhDong, item.moTa] for item in self.quyen_repo.list_all(limit=10000)],
        )

    def list_quy_phep(self, page: int = 1, size: int = 20) -> list[QuyPhep]:
        return self.quy_phep_repo.list_paginated(page, size)

    def create_quy_phep(self, payload: QuyPhepCreate) -> QuyPhep:
        data = payload.model_dump()
        data["ngayCapNhat"] = data["ngayCapNhat"] or datetime.now()
        return self._create_catalog_item(self.quy_phep_repo, QuyPhep(**data), payload.id_QuyPhep)

    def update_quy_phep(self, id_quy_phep: str, payload: QuyPhepUpdate) -> QuyPhep:
        data = payload.model_dump(exclude_unset=True)
        if "ngayCapNhat" not in data:
            data["ngayCapNhat"] = datetime.now()
        return self._update_catalog_item(self.quy_phep_repo, id_quy_phep, data)

    def delete_quy_phep(self, id_quy_phep: str) -> dict[str, str]:
        return self._delete_catalog_item(self.quy_phep_repo, id_quy_phep)

    async def import_quy_phep(self, file: UploadFile) -> dict[str, object]:
        return await self._import_catalog(
            file=file,
            factory=lambda row: QuyPhepCreate(
                id_QuyPhep=str(row[0]),
                id_NhanVien=str(row[1]),
                nam=int(row[2]),
                tongQuyPhep=Decimal(str(row[3])),
                soNgayDaDung=Decimal(str(row[4] if row[4] is not None else 0)),
                soNgayChoDuyet=Decimal(str(row[5] if row[5] is not None else 0)),
                ngayCapNhat=row[6] if row[6] is not None else None,
                trangThai=int(row[7] if row[7] is not None else 1),
            ),
            creator=self.create_quy_phep,
        )

    def export_quy_phep(self) -> StreamingResponse:
        return self._export_catalog(
            filename="quy_phep.xlsx",
            sheet_name="QuyPhep",
            headers=[
                "id_QuyPhep",
                "id_NhanVien",
                "nam",
                "tongQuyPhep",
                "soNgayDaDung",
                "soNgayChoDuyet",
                "ngayCapNhat",
                "trangThai",
            ],
            rows=[
                [
                    item.id_QuyPhep,
                    item.id_NhanVien,
                    item.nam,
                    item.tongQuyPhep,
                    item.soNgayDaDung,
                    item.soNgayChoDuyet,
                    item.ngayCapNhat,
                    item.trangThai,
                ]
                for item in self.quy_phep_repo.list_all(limit=10000)
            ],
        )

    def list_lich_hop(self, current_user: TaiKhoan) -> list[LichHopResponse]:
        if current_user.id_VaiTro in (ADMIN_ROLE_IDS | HCNS_ROLE_IDS):
            rows = self.lich_hop_repo.list_all()
        else:
            rows = self.lich_hop_repo.list_by_user(current_user.id_TaiKhoan)
        members_by_meeting = self.lich_hop_repo.list_members_by_meeting_ids([row.id_LichHop for row in rows])
        return [self._to_lich_hop_response(row, members_by_meeting.get(row.id_LichHop, [])) for row in rows]

    def list_nhan_vien_options(self) -> list[NhanVienOptionResponse]:
        rows = self.lich_hop_repo.list_employee_options()
        return [
            NhanVienOptionResponse(
                id_NhanVien=item.id_NhanVien,
                hoTen=item.hoTen,
                email=item.email,
                chucVu=item.chucVu,
            )
            for item in rows
        ]

    def tao_lich_hop(self, payload: LichHopCreate, current_user: TaiKhoan) -> LichHopResponse:
        phong = self.phong_hop_repo.get(payload.id_Phong)
        if phong is None:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Khong tim thay phong hop")
        if phong.trangThai != 1:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Phong hop dang ngung hoat dong")
        thoi_gian_bat_dau = self._normalize_meeting_time(payload.thoiGianBatDau)
        thoi_gian_ket_thuc = self._normalize_meeting_time(payload.thoiGianKetThuc)
        if thoi_gian_ket_thuc <= thoi_gian_bat_dau:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Thoi gian ket thuc phai sau thoi gian bat dau")
        if self.lich_hop_repo.check_conflict(payload.id_Phong, thoi_gian_bat_dau, thoi_gian_ket_thuc):
            raise HTTPException(status.HTTP_409_CONFLICT, "Phong hop da co lich vao khung gio nay")
        invalid_members = [
            id_nhan_vien
            for id_nhan_vien in set(payload.id_NhanVienThamGia)
            if not self.lich_hop_repo.employee_exists(id_nhan_vien)
        ]
        if invalid_members:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Nhan vien khong ton tai: {', '.join(invalid_members)}")

        lich_hop = LichHop(
            id_LichHop=f"LH-{uuid4().hex[:12]}",
            id_NhanVien=current_user.id_TaiKhoan,
            id_Phong=payload.id_Phong,
            tieuDe=payload.tieuDe,
            noiDung=payload.noiDung,
            thoiGianBatDau=thoi_gian_bat_dau,
            thoiGianKetThuc=thoi_gian_ket_thuc,
            mucDoUuTien=payload.mucDoUuTien,
            trangThai=0,
        )
        saved = self.lich_hop_repo.save(lich_hop)
        member_ids = [current_user.id_TaiKhoan, *payload.id_NhanVienThamGia]
        unique_member_ids = list(dict.fromkeys(member_ids))
        members = [
            ThanhVienLichHop(
                id_LichHop=saved.id_LichHop,
                id_NhanVien=id_nhan_vien,
                vaiTroThamGia="chu_tri" if id_nhan_vien == current_user.id_TaiKhoan else "tham_du",
                trangThaiThamGia="da_xac_nhan" if id_nhan_vien == current_user.id_TaiKhoan else "cho_xac_nhan",
            )
            for id_nhan_vien in unique_member_ids
        ]
        self.lich_hop_repo.replace_members(saved.id_LichHop, members)
        self._send_meeting_notifications(
            saved,
            members,
            title="Bạn có lịch họp mới đang chờ duyệt",
            prefix="Bạn được thêm vào lịch họp",
        )
        return self._to_lich_hop_response(saved, members)

    def huy_lich_hop(self, id_lich_hop: str, current_user: TaiKhoan) -> dict[str, str]:
        lich_hop = self.lich_hop_repo.get(id_lich_hop)
        if lich_hop is None:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Khong tim thay lich hop")
        is_admin = current_user.id_VaiTro in (ADMIN_ROLE_IDS | HCNS_ROLE_IDS)
        if not is_admin and lich_hop.id_NhanVien != current_user.id_TaiKhoan:
            raise HTTPException(status.HTTP_403_FORBIDDEN, "Ban khong co quyen huy lich hop nay")
        if lich_hop.trangThai == 2:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Lich hop da bi huy truoc do")
        lich_hop.trangThai = 2
        self.lich_hop_repo.save(lich_hop)
        return {"id_LichHop": id_lich_hop}

    def duyet_lich_hop(self, id_lich_hop: str, current_user: TaiKhoan) -> LichHopResponse:
        self._assert_can_moderate_meeting(current_user)
        lich_hop = self._get_lich_hop(id_lich_hop)
        if lich_hop.trangThai != 0:
            raise HTTPException(status.HTTP_409_CONFLICT, "Lịch họp không còn ở trạng thái chờ duyệt")
        conflicts = self.lich_hop_repo.list_conflicts(
            lich_hop.id_Phong,
            lich_hop.thoiGianBatDau,
            lich_hop.thoiGianKetThuc,
            exclude_id=lich_hop.id_LichHop,
            statuses={1},
        )
        if conflicts:
            raise HTTPException(status.HTTP_409_CONFLICT, "Phòng họp đã có lịch được duyệt trong khung giờ này")
        lich_hop.trangThai = 1
        saved = self.lich_hop_repo.save(lich_hop)
        self._send_meeting_notifications(
            saved,
            self.lich_hop_repo.list_members(saved.id_LichHop),
            title="Lịch họp đã được duyệt",
            prefix="Lịch họp đã được duyệt",
        )
        return self._to_lich_hop_response(saved)

    def tu_choi_lich_hop(
        self,
        id_lich_hop: str,
        payload: LichHopXuLyRequest,
        current_user: TaiKhoan,
    ) -> LichHopResponse:
        self._assert_can_moderate_meeting(current_user)
        if not payload.lyDo or not payload.lyDo.strip():
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Lý do từ chối là bắt buộc")
        lich_hop = self._get_lich_hop(id_lich_hop)
        if lich_hop.trangThai != 0:
            raise HTTPException(status.HTTP_409_CONFLICT, "Lịch họp không còn ở trạng thái chờ duyệt")
        lich_hop.trangThai = 3
        saved = self.lich_hop_repo.save(lich_hop)
        self._send_meeting_notifications(
            saved,
            self.lich_hop_repo.list_members(saved.id_LichHop),
            title="Lịch họp bị từ chối",
            prefix=f"Lịch họp bị từ chối. Lý do: {payload.lyDo.strip()}",
        )
        return self._to_lich_hop_response(saved)

    def can_thiep_lich_hop(
        self,
        id_lich_hop: str,
        payload: LichHopXuLyRequest,
        current_user: TaiKhoan,
    ) -> LichHopResponse:
        self._assert_can_moderate_meeting(current_user)
        lich_hop = self._get_lich_hop(id_lich_hop)
        if lich_hop.trangThai == 2:
            raise HTTPException(status.HTTP_409_CONFLICT, "Lịch họp đã bị hủy")

        conflicts = self.lich_hop_repo.list_conflicts(
            lich_hop.id_Phong,
            lich_hop.thoiGianBatDau,
            lich_hop.thoiGianKetThuc,
            exclude_id=lich_hop.id_LichHop,
            statuses={0, 1},
        )
        for item in conflicts:
            item.trangThai = 3
        lich_hop.trangThai = 1
        self.lich_hop_repo.save_many([*conflicts, lich_hop])
        for conflict in conflicts:
            self._send_meeting_notifications(
                conflict,
                self.lich_hop_repo.list_members(conflict.id_LichHop),
                title="Lịch họp bị thay đổi trạng thái",
                prefix="Lịch họp bị từ chối do trùng phòng với lịch ưu tiên cao hơn",
            )
        self._send_meeting_notifications(
            lich_hop,
            self.lich_hop_repo.list_members(lich_hop.id_LichHop),
            title="Lịch họp đã được duyệt",
            prefix="Lịch họp đã được duyệt sau khi xử lý xung đột phòng họp",
        )
        refreshed = self._get_lich_hop(id_lich_hop)
        return self._to_lich_hop_response(refreshed)

    def _create_catalog_item(self, repo: GenericDanhMucRepository, item, item_id: str):
        if repo.get(item_id):
            raise HTTPException(status.HTTP_409_CONFLICT, "Ma danh muc da ton tai")
        try:
            return repo.create(item)
        except IntegrityError:
            raise HTTPException(status.HTTP_409_CONFLICT, "Du lieu danh muc bi trung hoac dang tham chieu ban ghi khac")

    def _update_catalog_item(self, repo: GenericDanhMucRepository, item_id: str, payload: BaseModel | dict):
        item = repo.get(item_id)
        if item is None:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Khong tim thay danh muc")
        data = payload if isinstance(payload, dict) else payload.model_dump(exclude_unset=True)
        for key, value in data.items():
            setattr(item, key, value)
        try:
            return repo.save(item)
        except IntegrityError:
            raise HTTPException(status.HTTP_409_CONFLICT, "Du lieu danh muc bi trung hoac dang tham chieu ban ghi khac")

    def _delete_catalog_item(self, repo: GenericDanhMucRepository, item_id: str) -> dict[str, str]:
        item = repo.get(item_id)
        if item is None:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Khong tim thay danh muc")
        try:
            repo.delete(item)
        except IntegrityError:
            raise HTTPException(status.HTTP_409_CONFLICT, "Khong the xoa danh muc dang co du lieu lien ket")
        return {"id": item_id}

    async def _import_catalog(self, file: UploadFile, factory, creator) -> dict[str, object]:
        if not file.filename or not file.filename.lower().endswith(".xlsx"):
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Chi ho tro file Excel .xlsx")
        workbook = load_workbook(BytesIO(await file.read()))
        sheet = workbook.active
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        errors: list[dict[str, object]] = []
        success = 0
        for index, row in enumerate(rows, start=2):
            if not any(value is not None for value in row):
                continue
            try:
                creator(factory(row))
                success += 1
            except Exception as exc:
                errors.append({"dong": index, "loi": str(exc)})
        return {
            "tong_dong": len(rows),
            "thanh_cong": success,
            "that_bai": len(errors),
            "chi_tiet_loi": errors,
        }

    def _export_catalog(self, filename: str, sheet_name: str, headers: list[str], rows: list[list[object]]) -> StreamingResponse:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    def _get_phong_hop(self, id_phong: str) -> PhongHop:
        phong = self.phong_hop_repo.get(id_phong)
        if phong is None:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Khong tim thay phong hop")
        return phong

    def _get_lich_hop(self, id_lich_hop: str) -> LichHop:
        lich_hop = self.lich_hop_repo.get(id_lich_hop)
        if lich_hop is None:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy lịch họp")
        return lich_hop

    def _assert_can_moderate_meeting(self, current_user: TaiKhoan) -> None:
        if current_user.id_VaiTro not in MEETING_MODERATOR_ROLE_IDS:
            raise HTTPException(status.HTTP_403_FORBIDDEN, "Bạn không có quyền xử lý lịch họp")

    def _normalize_meeting_time(self, value: datetime) -> datetime:
        if value.tzinfo is None:
            return value
        return value.astimezone(APP_TIMEZONE).replace(tzinfo=None)

    def _format_meeting_time(self, value: datetime) -> str:
        local_value = self._normalize_meeting_time(value)
        return local_value.strftime("%H:%M ngày %d/%m/%Y")

    def _send_meeting_notifications(
        self,
        lich_hop: LichHop,
        members: list[ThanhVienLichHop],
        title: str,
        prefix: str,
    ) -> None:
        receiver_ids = sorted({member.id_NhanVien for member in members if member.id_NhanVien})
        if not receiver_ids:
            return
        phong = self.phong_hop_repo.get(lich_hop.id_Phong)
        room_name = phong.tenPhong if phong else lich_hop.id_Phong
        content = (
            f"{prefix}: {lich_hop.tieuDe}. "
            f"Phòng: {room_name}. "
            f"Thời gian: {self._format_meeting_time(lich_hop.thoiGianBatDau)} - "
            f"{self._format_meeting_time(lich_hop.thoiGianKetThuc)}."
        )
        now = datetime.now()
        try:
            for receiver_id in receiver_ids:
                self.lich_hop_repo.session.add(
                    ThongBao(
                        id_ThongBao=f"TB-{uuid4().hex[:12]}",
                        id_NguoiNhan=receiver_id,
                        tieuDe=title,
                        noiDung=content,
                        loaiThongBao="LICH_HOP",
                        trangThaiDoc=0,
                        thoiGianGui=now,
                    )
                )
            self.lich_hop_repo.session.commit()
        except Exception:
            self.lich_hop_repo.session.rollback()
            raise

    def _to_lich_hop_response(self, item: LichHop, members: list[ThanhVienLichHop] | None = None) -> LichHopResponse:
        return LichHopResponse(
            id_LichHop=item.id_LichHop,
            id_NhanVien=item.id_NhanVien,
            id_Phong=item.id_Phong,
            tieuDe=item.tieuDe,
            noiDung=item.noiDung,
            thoiGianBatDau=item.thoiGianBatDau,
            thoiGianKetThuc=item.thoiGianKetThuc,
            mucDoUuTien=item.mucDoUuTien,
            trangThai=item.trangThai,
            thanhVien=[
                ThanhVienLichHopResponse(
                    id_LichHop=member.id_LichHop,
                    id_NhanVien=member.id_NhanVien,
                    vaiTroThamGia=member.vaiTroThamGia,
                    trangThaiThamGia=member.trangThaiThamGia,
                )
                for member in (members or self.lich_hop_repo.list_members(item.id_LichHop))
            ],
        )
