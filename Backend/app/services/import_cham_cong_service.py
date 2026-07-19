import csv
from datetime import datetime
from decimal import Decimal
from io import BytesIO, StringIO
from uuid import uuid4

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlmodel import Session

from app.models.bang_cong import BangCong
from app.repositories.bang_cong_repo import BangCongRepository
from app.schemas.cham_cong_schema import BangCongPreviewItem, ImportPreviewResponse
from app.utils.tinh_cong_redmine import tinh_cong_redmine
from app.utils.tinh_cong_tingop import tinh_cong_tingop

IMPORT_PREVIEW_CACHE: dict[str, list[BangCongPreviewItem]] = {}
MAX_IMPORT_SIZE = 10 * 1024 * 1024


class ImportChamCongService:
    def __init__(self, session: Session) -> None:
        self.repo = BangCongRepository(session)

    async def preview_import(self, file: UploadFile, nguon: str) -> ImportPreviewResponse:
        """Read attendance export, validate template, calculate preview, and cache it temporarily."""
        if nguon not in {"tingop", "redmine"}:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Nguồn import phải là tingop hoặc redmine")
        content = await file.read()
        if len(content) > MAX_IMPORT_SIZE:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "File import vượt quá giới hạn 10MB")
        filename = file.filename or ""
        if not filename.lower().endswith((".xlsx", ".csv")):
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Chỉ hỗ trợ file .xlsx hoặc .csv")

        rows = self._read_rows(filename, content)
        self._validate_template(rows, nguon)
        if nguon == "tingop":
            preview_rows = tinh_cong_tingop(rows)
        else:
            preview_rows = tinh_cong_redmine(rows)
        items = [BangCongPreviewItem(**row) for row in preview_rows]
        preview_id = f"PREVIEW-{uuid4().hex[:12]}"
        IMPORT_PREVIEW_CACHE[preview_id] = items
        return ImportPreviewResponse(preview_id=preview_id, tong_dong=len(items), du_lieu=items)

    def confirm_import(self, preview_id: str) -> dict[str, int]:
        """Persist cached preview rows into BangCong with upsert behavior."""
        items = IMPORT_PREVIEW_CACHE.pop(preview_id, None)
        if not items:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Không tìm thấy dữ liệu preview import")
        inserted = 0
        updated = 0
        for item in items:
            # BR18-3: import lại cùng id_NhanVien + tenBangCong + tuNgay/denNgay thì ghi đè, không tạo trùng.
            existing = self.repo.find_existing(
                item.id_NhanVien,
                item.tenBangCong,
                item.tuNgay,
                item.denNgay,
            )
            if existing:
                existing.loaiHinhTinhCong = item.loaiHinhTinhCong
                existing.tongGioLogtime = item.tongGioLogtime
                existing.tongGioLogtimeThucTe = item.tongGioLogtimeThucTe
                existing.tenDuAn_Task = item.tenDuAn_Task
                existing.soLanDiMuon = item.soLanDiMuon
                existing.ngayCapNhat = datetime.now()
                existing.trangThai = 0
                self.repo.save(existing)
                updated += 1
            else:
                bang_cong = BangCong(
                    id_BangCong=f"BC-{uuid4().hex[:12]}",
                    id_NhanVien=item.id_NhanVien,
                    tenBangCong=item.tenBangCong,
                    loaiHinhTinhCong=item.loaiHinhTinhCong,
                    tongGioLogtime=item.tongGioLogtime,
                    tongGioLogtimeThucTe=item.tongGioLogtimeThucTe,
                    tenDuAn_Task=item.tenDuAn_Task,
                    soLanDiMuon=item.soLanDiMuon,
                    tuNgay=item.tuNgay,
                    denNgay=item.denNgay,
                    ngayCapNhat=datetime.now(),
                    trangThai=0,
                )
                self.repo.save(bang_cong)
                inserted += 1
        return {"them_moi": inserted, "cap_nhat": updated}

    def _read_rows(self, filename: str, content: bytes) -> list[dict[str, object]]:
        if filename.lower().endswith(".csv"):
            text = content.decode("utf-8-sig")
            return list(csv.DictReader(StringIO(text)))
        workbook = load_workbook(BytesIO(content), data_only=True)
        sheet = workbook.active
        headers = [str(cell.value).strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        return [
            {headers[index]: cell for index, cell in enumerate(row)}
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if any(cell is not None for cell in row)
        ]

    def _validate_template(self, rows: list[dict[str, object]], nguon: str) -> None:
        if not rows:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "File import không có dữ liệu")
        required = {"id_NhanVien", "tenBangCong", "ngay"} if nguon == "tingop" else {
            "id_NhanVien",
            "tenBangCong",
            "tuNgay",
            "denNgay",
            "tongGioLogtime",
        }
        if not required.issubset(set(rows[0].keys())):
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "File import sai định dạng template")
