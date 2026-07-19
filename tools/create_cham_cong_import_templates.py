from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path("docs")


def style_sheet(workbook: Workbook, title: str) -> None:
    sheet = workbook.active
    sheet.title = title
    header_fill = PatternFill("solid", fgColor="4338CA")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = min(max(max_length + 2, 14), 36)
    sheet.freeze_panes = "A2"


def save_tingop_template() -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["id_NhanVien", "tenBangCong", "ngay", "diMuon"])
    sheet.append(["NV001", "Chấm công nội bộ tháng 7/2026", "2026-07-01", "false"])
    sheet.append(["NV001", "Chấm công nội bộ tháng 7/2026", "2026-07-02", "true"])
    sheet.append(["NV002", "Chấm công nội bộ tháng 7/2026", "2026-07-01", "false"])
    sheet.append(["NV003", "Chấm công nội bộ tháng 7/2026", "2026-07-01", "false"])
    style_sheet(workbook, "TingOp")
    output_path = OUTPUT_DIR / "mau_import_cham_cong_tingop.xlsx"
    workbook.save(output_path)
    return output_path


def save_redmine_template() -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["id_NhanVien", "tenBangCong", "tuNgay", "denNgay", "tongGioLogtime", "tenDuAn_Task"])
    sheet.append(["NV001", "Bảng công tháng 7/2026 - Dự án A", "2026-07-01", "2026-07-28", 160, "Project A"])
    sheet.append(["NV002", "Bảng công tháng 7/2026 - Dự án A", "2026-07-01", "2026-07-28", 152, "Project A"])
    sheet.append(["NV003", "Bảng công tháng 7/2026 - Dự án B", "2026-07-01", "2026-07-28", 144, "Project B"])
    style_sheet(workbook, "Redmine")
    output_path = OUTPUT_DIR / "mau_import_cham_cong_redmine.xlsx"
    workbook.save(output_path)
    return output_path


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    outputs = [save_tingop_template(), save_redmine_template()]
    for output in outputs:
        print(output)


if __name__ == "__main__":
    main()
